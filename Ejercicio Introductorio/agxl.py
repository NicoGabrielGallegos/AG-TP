# -*- coding: latin-1 -*-

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
from openpyxl.utils import get_column_letter

# Constantes de alineamientos, fuentes, colores, etc
alineamiento_centro_centro = Alignment(horizontal='center', vertical='center')

# Formato general
def general_formateo(ws) -> None:
    ws.row_dimensions[1].height = 15
    ws.column_dimensions['A'].width = 2.85

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↓↓↓ \ Hoja Generaciones \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #
# Formateo de la hoja de generaciones
def generaciones_formateo(ws, ciclos: int) -> None:
    general_formateo(ws)
    for ciclo in range(ciclos):
        insertion_column = 2 + (ciclo) * 6
        for relative_column in range(5):
            absolute_column = insertion_column + relative_column
            ws.column_dimensions[get_column_letter(absolute_column)].width = (10, 12, 33, 13, 50)[relative_column] + 0.71
        ws[f'{get_column_letter(insertion_column)}2'].font = Font(size=16)
        ws.cell(row=2, column=insertion_column, value=f"Ciclo {ciclo + 1}").alignment = alineamiento_centro_centro
        ws.merge_cells(start_row=2, end_row=2, start_column=insertion_column, end_column=(insertion_column+4))
    ws.row_dimensions[2].height = 23
    return

# Insertar la tabla de una generación
def generaciones_insertar_tabla(ws, ciclo: int, generacion: int, poblacion: list[int], evaluaciones: list[float]) -> None:
    insertion_row = 4 + (generacion - 1) * 13
    shift_column = (ciclo) * 6
    ws.cell(row=insertion_row, column=(2 + shift_column), value=f"Generacion {generacion}").alignment = alineamiento_centro_centro
    ws.merge_cells(start_row=insertion_row, end_row=insertion_row, start_column=(2 + shift_column), end_column=(6 + shift_column))

    for relative_column in range(4):
        absolute_column = 2 + relative_column + shift_column
        ws.cell(row=insertion_row + 1, column=absolute_column, value=("Individuo", "Cromosoma", "Cromosoma (binario)", "Fitness")[relative_column]).alignment = alineamiento_centro_centro
        
    ws.row_dimensions[insertion_row].height = ws.row_dimensions[insertion_row + 1].height = 23
    for relative_row in range(10):
        absolute_row = 2 + insertion_row + relative_row
        ws.row_dimensions[absolute_row].height = 23
        ws.cell(row=absolute_row, column=(2 + shift_column), value=(str(relative_row + 1) + ".")).alignment = alineamiento_centro_centro
        ws.cell(row=absolute_row, column=(3 + shift_column), value = poblacion[relative_row]).alignment = alineamiento_centro_centro
        ws.cell(row=absolute_row, column=(4 + shift_column), value = format(poblacion[relative_row], "030b")).alignment = alineamiento_centro_centro
        ws.cell(row=absolute_row, column=(5 + shift_column), value = evaluaciones[relative_row]).alignment = alineamiento_centro_centro
    return

# Insertar la gráfica de una generación
def generaciones_insertar_grafica(ws, ciclo: int, generacion: int) -> None:
    insertion_row = 4 + (generacion - 1) * 13
    shift_column = (ciclo) * 6
    grafica = PieChart()
    encabezados = Reference(ws, min_col=(2 + shift_column), min_row=(insertion_row + 4), max_row=(insertion_row + 13))
    datos = Reference(ws, min_col=(3 + shift_column), min_row=(insertion_row + 4), max_row=(insertion_row + 13))
    grafica.add_data(datos, titles_from_data=False)
    grafica.set_categories(encabezados)
    grafica.legend = None
    grafica.dataLabels = DataLabelList()
    grafica.dataLabels.showPercent = True
    grafica.dataLabels.showCatName = True
    grafica.dataLabels.position = "inEnd"
    grafica.style = 10

    anclaje = TwoCellAnchor()
    anclaje._from.col = (5 + shift_column)
    anclaje._from.row = insertion_row
    anclaje.to.col = (6 + shift_column)
    anclaje.to.row = insertion_row + 11
    grafica.anchor = anclaje

    ws.add_chart(grafica)
    return
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↑↑↑ \ Hoja Generaciones \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↓↓↓ \ Hoja Ciclos \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #
# Formateo de la hoja de ciclos
def ciclos_formateo(ws, generaciones) -> None:
    general_formateo(ws)
    for row in range(11 + generaciones):
        ws.row_dimensions[row + 2].height = 23

# Insertar la tabla de un ciclo
def ciclos_insertar_tabla(ws, ciclo: int, generaciones: int, cantidad_individuos: int) -> None:
    insertion_column = 2 + (ciclo) * 5
    shift_column = (ciclo) * 6
    ws.cell(row=2, column=insertion_column, value=f"Ciclo {ciclo + 1}").alignment = alineamiento_centro_centro
    ws.merge_cells(start_row=2, end_row=2, start_column=insertion_column, end_column=(insertion_column + 3))

    for relative_column in range(4):
        absolute_column = insertion_column + relative_column
        ws.cell(row=12, column=absolute_column, value=("Generacion", "Maximo", "Minimo", "Promedio")[relative_column]).alignment = alineamiento_centro_centro
        ws.column_dimensions[get_column_letter(absolute_column)].width = 20.71

    for relative_row in range(generaciones):
        absolute_row = 13 + relative_row
        ws.cell(row=absolute_row, column=insertion_column, value=(str(relative_row + 1) + ".")).alignment = alineamiento_centro_centro
        ws.cell(row=absolute_row, column=(insertion_column + 1), value=f"=MAX(Generaciones!{get_column_letter(3 + shift_column)}{6 + relative_row * 13}:{get_column_letter(3 + shift_column)}{15 + relative_row * 13})").alignment = alineamiento_centro_centro
        ws.cell(row=absolute_row, column=(insertion_column + 2), value=f"=MIN(Generaciones!{get_column_letter(3 + shift_column)}{6 + relative_row * 13}:{get_column_letter(3 + shift_column)}{15 + relative_row * 13})").alignment = alineamiento_centro_centro
        ws.cell(row=absolute_row, column=(insertion_column + 3), value=f"=SUM(Generaciones!{get_column_letter(3 + shift_column)}{6 + relative_row * 13}:{get_column_letter(3 + shift_column)}{15 + relative_row * 13})/{cantidad_individuos}").alignment = alineamiento_centro_centro
    return

# Insertar la gráfica de un ciclo
def ciclos_insertar_grafica(ws, ciclo: int, generaciones: int):
    insertion_column = 2 + (ciclo) * 5
    grafica = LineChart()
    grafica.title = None
    grafica.style = 10
    grafica.y_axis.title = None
    grafica.x_axis.title = None
    datos = Reference(ws, min_col=(insertion_column + 1), max_col=(insertion_column + 3), min_row=12, max_row=(12 + generaciones))
    grafica.add_data(datos, titles_from_data=True)
    grafica.legend.position = 't'

    for serie in grafica.series:
        serie.marker.symbol = "circle"
        serie.graphicalProperties.line.width = 20000
        #serie.graphicalProperties.line.noFill = True

    anclaje = TwoCellAnchor()
    anclaje._from.col = insertion_column - 1
    anclaje._from.row = 2
    anclaje.to.col = insertion_column + 3
    anclaje.to.row = 11
    grafica.anchor = anclaje

    ws.add_chart(grafica)
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↑↑↑ \ Hoja Ciclos \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #