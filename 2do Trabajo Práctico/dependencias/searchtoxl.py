from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.cell.cell import Cell
from dependencias.knapsack import BagItem, Bag


# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↓↓↓ \ Variables \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #

# Columnas utilizadas en las hojas
_columns = ['B', 'C', 'D']

# Alineamiento de celdas
_center = Alignment(horizontal='center', vertical='center')
_center_v = Alignment(vertical='center')

# Bordes de las tablas
_border_out: Side = Side(border_style="thin", color="595959")
_border_in_v: Side = Side(border_style="hair", color="D9D9D9")
_border_in_h: Side = Side(border_style="thin", color="A6A6A6")

# Filas de las tablas
#_row_

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↑↑↑ \ Variables \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↓↓↓ \ General \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #

def format(wb: Workbook, items: list[BagItem], type: str = "Exhaustive") -> None:

    # Creación de las hojas "Objetos" y "Soluciones"
    wb.active.title = "Objetos"
    wb.create_sheet("Soluciones")

    # Formateo de la hoja "Objetos"
    ws_items: Worksheet = wb["Objetos"]
    format_items(ws_items, len(items) + 3, type)
    for i in range(len(items)):
        add_item(ws_items, i + 3, items[i])

    if type == "Greedy":
        _columns.pop()

    # Formateo de la hoja "Soluciones"
    ws_nodes: Worksheet = wb["Soluciones"]
    if type == "Exhaustive":
        format_nodes(ws_nodes, pow(2,len(items)) + 4)
    else:
        format_nodes(ws_nodes, 2)

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↑↑↑ \ General \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↓↓↓ \ Hoja Objetos \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #

def format_items(ws: Worksheet, rows: int, type: str = "Exhaustive") -> None:

    headers = {'B': "Objeto", 'C': "Peso", 'D': "Valor", 'E': "Peso/Valor"}
    if type == "Greedy":
        _columns.append('E')

    ws.column_dimensions['A'].width = 2.85

    # Encabezado
    for col in _columns:
        ws.column_dimensions[col].width = 13.71
        cell: Cell = ws[f'{col}2']
        cell.font = Font(name="Consolas", color="F2F2F2")
        cell.fill = PatternFill(fgColor="262626", fill_type="solid")
        cell.value = headers[col]
        cell.alignment = _center
        cell.border = Border(_border_out, _border_out, _border_out, _border_out)
    ws.row_dimensions[2].height = 24

    # Objetos
    for row in range(3, rows):
        ws.row_dimensions[row].height = 18
        if row % 2 == 0:
            COLOR = {'B': "31869B", 'C': "60497A", 'D': "76933C", 'E': "963634"}
            FILL = PatternFill(fgColor="D9D9D9", fill_type="solid")
        else:
            COLOR = {'B': "4BACC6", 'C': "8064A2", 'D': "9BBB59", 'E': "C0504D"}
            FILL = PatternFill(fgColor="FFFFFF", fill_type="solid")

        for col in _columns:
            cell: Cell = ws[f'{col}{row}']
            cell.font = Font(name="Consolas", color=COLOR[col])
            cell.fill = FILL
            cell.alignment = _center

            if col == _columns[0]:
                cell.border = Border(bottom=_border_in_h, left=_border_out)
            elif col == _columns[-1]:
                cell.border = Border(bottom=_border_in_h, right=_border_out)
            else:
                cell.border = Border(bottom=_border_in_h, left=_border_in_v, right=_border_in_v)
    
    for col in _columns:
        cell: Cell = ws[f'{col}{rows-1}']
        if col == _columns[0]:
            cell.border = Border(bottom=_border_out, left=_border_out)
        elif col == _columns[-1]:
            cell.border = Border(bottom=_border_out, right=_border_out)
        else:
            cell.border = Border(bottom=_border_out, left=_border_in_v, right=_border_in_v)

def add_item(ws: Worksheet, row: int, item: BagItem) -> None:

    values = { 'B': row-3, 'C': item.weight, 'D': item.value, 'E': item.value_weight_ratio}

    for col in _columns:
        cell: Cell = ws[f'{col}{row}']
        cell.value = values[col]

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↑↑↑ \ Hoja Objetos \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↓↓↓ \ Hoja Soluciones \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #


def format_nodes(ws: Worksheet, rows: int) -> None:
    headers = {'B': "Peso", 'C': "Valor", 'D': "Contenido", 2: "Posibles soluciones", rows: "Soluciones óptimas"}

    for row in [2, 3, rows, rows+1]:
        ws.row_dimensions[row].height = 24
        ws.row_dimensions[row].height = 24
    ws.column_dimensions['A'].width = 2.85

    # Encabezados
    for row in [2, rows]:
        ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=4)
        cell: Cell = ws[f'B{row}']
        cell.font = Font(name="Consolas", color="F2F2F2")
        cell.fill = PatternFill(fgColor="262626", fill_type="solid")
        cell.value = headers[row]
        cell.alignment = _center
        cell.border = Border(_border_out, _border_out, _border_out, _border_out)
    
    for col in _columns:
        ws.column_dimensions[col].width = 13.71
        for row in [3, rows+1]:
            cell: Cell = ws[f'{col}{row}']
            cell.font = Font(name="Consolas", color="F2F2F2")
            cell.fill = PatternFill(fgColor="262626", fill_type="solid")
            cell.value = headers[col]
            cell.alignment = _center
            cell.border = Border(_border_out, _border_out, _border_out, _border_out)
    ws.row_dimensions[2].height = 24

    # Ancho de la 3era columna
    width: str = len("Mochila(1, 2, 3, 4, 5, 6, 7, 8, 9, 10)")

    ws.column_dimensions["D"].width = 1.2 * width + 1.71

    # Soluciones
    for row in range(4, rows):
        ws.row_dimensions[row].height = 18
        if row % 2 == 1:
            FILL = PatternFill(fgColor="D9D9D9", fill_type="solid")
        else:
            FILL = PatternFill(fgColor="FFFFFF", fill_type="solid")

        for col in _columns:
            cell: Cell = ws[f'{col}{row}']
            cell.fill = FILL
            if col != "D":
                cell.alignment = _center
            else:
                cell.alignment = _center_v

            if col == _columns[0]:
                cell.border = Border(bottom=_border_in_h, left=_border_out)
            elif col == _columns[-1]:
                cell.border = Border(bottom=_border_in_h, right=_border_out)
            else:
                cell.border = Border(bottom=_border_in_h, left=_border_in_v, right=_border_in_v)
    
    for col in _columns:
        cell: Cell = ws[f'{col}{rows-1}']
        if col == _columns[0]:
            cell.border = Border(bottom=_border_out, left=_border_out)
        elif col == _columns[-1]:
            cell.border = Border(bottom=_border_out, right=_border_out)
        else:
            cell.border = Border(bottom=_border_out, left=_border_in_v, right=_border_in_v)

def add_node(ws: Worksheet, row: int, bag: Bag) -> None:
    value = {'B': bag.total_weight, 'C': bag.total_value, 'D': bag.__str__(False)}
    row += 4

    if bag.is_valid:
        if row % 2 == 0:
            COLOR = {'B': "8064A2", 'C': "9BBB59", 'D': "4BACC6"}
        else:
            COLOR = {'B': "60497A", 'C': "76933C", 'D': "31869B"}
    else:
        if row % 2 == 0:
            COLOR = {'B': "808080", 'C': "808080", 'D': "808080"}
        else:
            COLOR = {'B': "505050", 'C': "505050", 'D': "505050"}
    
    for col in _columns:
        cell: Cell = ws[f'{col}{row}']
        cell.font = Font(name="Consolas", color=COLOR[col])
        cell.value = value[col]



def add_optimum_node(ws: Worksheet, q_items: int, bags: list[Bag], type: str = "Exhaustive") -> None:
    if type == "Exhaustive":
        row_start = 6 + pow(2, q_items)
    else:
        row_start = 4
    i: int = -1

    for bag in bags:
        i += 1

        values = {'B': bag.total_weight, 'C': bag.total_value, 'D': bag.__str__(False)}
        row = row_start + i

        if i % 2 == 0:
            COLOR = {'B': "8064A2", 'C': "9BBB59", 'D': "4BACC6"}
            FILL = PatternFill(fgColor="FFFFFF", fill_type="solid")
        else:
            COLOR = {'B': "60497A", 'C': "76933C", 'D': "31869B"}
            FILL = PatternFill(fgColor="D9D9D9", fill_type="solid")

        for col in _columns:
            cell: Cell = ws[f'{col}{row}']
            cell.font = Font(name="Consolas", color=COLOR[col])
            cell.fill = FILL
            cell.value = values[col]
            if col != "D":
                cell.alignment = _center
            else:
                cell.alignment = _center_v
            if col == _columns[0]:
                cell.border = Border(bottom=_border_in_h, left=_border_out)
            elif col == _columns[-1]:
                cell.border = Border(bottom=_border_in_h, right=_border_out)
            else:
                cell.border = Border(bottom=_border_in_h, left=_border_in_v, right=_border_in_v)
    
    for col in _columns:
        cell: Cell = ws[f'{col}{row_start + i}']
        if col == _columns[0]:
            cell.border = Border(bottom=_border_out, left=_border_out)
        elif col == _columns[-1]:
            cell.border = Border(bottom=_border_out, right=_border_out)
        else:
            cell.border = Border(bottom=_border_out, left=_border_in_v, right=_border_in_v)

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- ↑↑↑ \ Hoja Soluciones \ -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=- #